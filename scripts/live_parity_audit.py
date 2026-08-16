from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin

import openpyxl
import requests
from bs4 import BeautifulSoup

ROOT = "https://na.org.za"
FEEDS = {
    "region": f"{ROOT}/wp-admin/admin-ajax.php?action=meetings",
    "johannesburg": f"{ROOT}/jhb/wp-admin/admin-ajax.php?action=meetings",
    "pretoria": f"{ROOT}/pta/wp-admin/admin-ajax.php?action=meetings",
    "western-cape": f"{ROOT}/wc/wp-admin/admin-ajax.php?action=meetings",
    "kwazulu-natal": f"{ROOT}/kzn/wp-admin/admin-ajax.php?action=meetings",
}
SITEMAPS = {
    "region": f"{ROOT}/sitemap_index.xml",
    "johannesburg": f"{ROOT}/jhb/sitemap_index.xml",
    "pretoria": f"{ROOT}/pta/sitemap_index.xml",
    "western-cape": f"{ROOT}/wc/sitemap_index.xml",
    "kwazulu-natal": f"{ROOT}/kzn/sitemap_index.xml",
}
WORKBOOK = Path("/home/ubuntu/upload/pasted_file_3nDiNL_NA-South-Africa-All-Meetings.xlsx")
OUT = Path("/home/ubuntu/na-south-africa-rebuild")
SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "NA-South-Africa-Parity-Audit/1.0"})


def get_json(url: str):
    response = SESSION.get(url, timeout=45)
    response.raise_for_status()
    return response.json()


def get_text(url: str):
    response = SESSION.get(url, timeout=45)
    response.raise_for_status()
    return response.text


def norm(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def feed_key(item):
    return (norm(item.get("name")), norm(item.get("day")), norm(item.get("time")), norm(item.get("attendance_option")), norm(item.get("region")))


def workbook_rows():
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    sheet = wb["All Meetings"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        rows.append(row)
    return rows


def sitemap_urls(index_url: str):
    xml = get_text(index_url)
    soup = BeautifulSoup(xml, "xml")
    child = [loc.get_text(strip=True) for loc in soup.find_all("loc")]
    urls = []
    if child and any("sitemap" in url for url in child):
        for child_url in child:
            child_xml = get_text(child_url)
            child_soup = BeautifulSoup(child_xml, "xml")
            urls.extend(loc.get_text(strip=True) for loc in child_soup.find_all("loc"))
    else:
        urls = child
    return sorted(set(urls))


def main():
    generated = datetime.now(timezone.utc).isoformat()
    feeds = {}
    errors = {}
    for site, url in FEEDS.items():
        try:
            data = get_json(url)
            feeds[site] = data if isinstance(data, list) else []
        except Exception as exc:
            errors[f"feed:{site}"] = repr(exc)
            feeds[site] = []

    all_feed_rows = [(site, row) for site, rows in feeds.items() for row in rows]
    unique_by_id = {}
    for site, row in all_feed_rows:
        key = (site, row.get("id"), row.get("slug"))
        unique_by_id[key] = row

    unique_by_fingerprint = defaultdict(list)
    for site, row in all_feed_rows:
        unique_by_fingerprint[feed_key(row)].append((site, row))

    workbook = workbook_rows()
    workbook_by_key = defaultdict(list)
    for row in workbook:
        workbook_by_key[(norm(row.get("Meeting Name")), norm(row.get("Day")), norm(row.get("Start")), norm(row.get("Attendance")), norm(row.get("Region")))].append(row)

    live_unique_candidates = []
    seen = set()
    for site, row in all_feed_rows:
        fp = feed_key(row)
        if fp not in seen:
            seen.add(fp)
            live_unique_candidates.append({"feed_site": site, **row})

    live_keys = set(feed_key(row) for row in live_unique_candidates)
    workbook_keys = set(workbook_by_key)
    live_missing_in_workbook = sorted(live_keys - workbook_keys)
    workbook_missing_live = sorted(workbook_keys - live_keys)

    sitemap_inventory = {}
    all_urls = set()
    for site, url in SITEMAPS.items():
        try:
            urls = sitemap_urls(url)
            sitemap_inventory[site] = urls
            all_urls.update(urls)
        except Exception as exc:
            errors[f"sitemap:{site}"] = repr(exc)
            sitemap_inventory[site] = []

    meeting_urls = sorted(url for url in all_urls if "/meetings/" in url or "/blog/meetings/" in url)
    page_urls = sorted(url for url in all_urls if url not in meeting_urls)
    online = [row for row in live_unique_candidates if norm(row.get("attendance_option")) == "online"]
    in_person = [row for row in live_unique_candidates if norm(row.get("attendance_option")) == "in_person"]
    hybrid = [row for row in live_unique_candidates if norm(row.get("attendance_option")) == "hybrid"]
    inactive = [row for row in live_unique_candidates if norm(row.get("attendance_option")) == "inactive"]

    report = {
        "generatedAt": generated,
        "root": ROOT,
        "feedUrls": FEEDS,
        "sitemapUrls": SITEMAPS,
        "errors": errors,
        "feedCounts": {site: len(rows) for site, rows in feeds.items()},
        "rawFeedCount": len(all_feed_rows),
        "uniqueFingerprintCount": len(live_unique_candidates),
        "uniqueIdSlugCount": len(unique_by_id),
        "liveAttendanceCounts": dict(Counter(norm(row.get("attendance_option")) for row in live_unique_candidates)),
        "liveRegionCounts": dict(Counter(norm(row.get("region")) for row in live_unique_candidates)),
        "workbookRowCount": len(workbook),
        "workbookAttendanceCounts": dict(Counter(norm(row.get("Attendance")) for row in workbook)),
        "workbookRegionCounts": dict(Counter(norm(row.get("Region")) for row in workbook)),
        "sitemapUrlCounts": {site: len(urls) for site, urls in sitemap_inventory.items()},
        "sitemapTotalUniqueUrls": len(all_urls),
        "sitemapMeetingLikeUrls": len(meeting_urls),
        "sitemapNonMeetingUrls": len(page_urls),
        "liveOnlineCount": len(online),
        "liveInPersonCount": len(in_person),
        "liveHybridCount": len(hybrid),
        "liveInactiveCount": len(inactive),
        "liveMissingInWorkbookCount": len(live_missing_in_workbook),
        "workbookMissingFromLiveCount": len(workbook_missing_live),
        "liveMissingInWorkbook": live_missing_in_workbook,
        "workbookMissingFromLive": workbook_missing_live,
        "sitemapInventory": sitemap_inventory,
        "liveUniqueRows": live_unique_candidates,
    }
    (OUT / "LIVE_PARITY_AUDIT_REPORT.json").write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    summary = [
        "# Live parity audit summary",
        "",
        f"Generated: {generated}",
        "",
        "| Measure | Result |",
        "|---|---:|",
        f"| Raw rows across five feeds | {len(all_feed_rows)} |",
        f"| Unique live fingerprints | {len(live_unique_candidates)} |",
        f"| Workbook rows | {len(workbook)} |",
        f"| Online | {len(online)} |",
        f"| In-person | {len(in_person)} |",
        f"| Hybrid | {len(hybrid)} |",
        f"| Inactive | {len(inactive)} |",
        f"| Sitemap URLs | {len(all_urls)} |",
        f"| Meeting-like sitemap URLs | {len(meeting_urls)} |",
        f"| Non-meeting sitemap URLs | {len(page_urls)} |",
        f"| Live fingerprints missing from workbook | {len(live_missing_in_workbook)} |",
        f"| Workbook fingerprints missing from live feeds | {len(workbook_missing_live)} |",
        "",
        "## Feed counts",
        "",
        "| Feed | Rows |",
        "|---|---:|",
    ]
    summary.extend(f"| {site} | {len(rows)} |" for site, rows in feeds.items())
    summary += ["", "## Errors", ""]
    summary.extend(f"- `{key}`: {value}" for key, value in errors.items()) if errors else summary.append("None.")
    (OUT / "LIVE_PARITY_AUDIT_SUMMARY.md").write_text("\n".join(summary) + "\n", encoding="utf-8")
    print(json.dumps({k: report[k] for k in ["feedCounts", "rawFeedCount", "uniqueFingerprintCount", "workbookRowCount", "liveAttendanceCounts", "sitemapUrlCounts", "sitemapTotalUniqueUrls", "liveMissingInWorkbookCount", "workbookMissingFromLiveCount", "errors"]}, indent=2))


if __name__ == "__main__":
    main()
