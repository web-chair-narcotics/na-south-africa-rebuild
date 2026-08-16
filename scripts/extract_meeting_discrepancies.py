from pathlib import Path
import json
import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook('/home/ubuntu/upload/pasted_file_3nDiNL_NA-South-Africa-All-Meetings.xlsx', read_only=True, data_only=True)
ws = wb['All Meetings']
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
rows = [dict(zip(headers, vals)) for vals in ws.iter_rows(min_row=2, values_only=True)]
report = json.loads(Path('/home/ubuntu/na-south-africa-rebuild/LIVE_PARITY_AUDIT_REPORT.json').read_text())
live = report['liveUniqueRows']

def n(v): return '' if v is None else ' '.join(str(v).strip().lower().split())
def slug(url): return str(url).rstrip('/').split('/')[-1] if url else ''
wb_by = defaultdict(list)
for r in rows: wb_by[slug(r.get('Meeting Page URL'))].append(r)
live_by = defaultdict(list)
for r in live: live_by[slug(r.get('url'))].append(r)
for name in sorted(set(wb_by) | set(live_by)):
    if (name not in wb_by) or (name not in live_by):
        print('\nSLUG', name)
        print('WORKBOOK', [{k:r.get(k) for k in ['Meeting ID','Meeting Name','Day','Start','Attendance','Region','Meeting Page URL']} for r in wb_by.get(name, [])])
        print('LIVE', [{k:r.get(k) for k in ['id','name','day','time','attendance_option','region','url','data_source_name','source_slug']} for r in live_by.get(name, [])])
    elif len(wb_by[name]) != len(live_by[name]):
        print('\nCOUNT MISMATCH', name, len(wb_by[name]), len(live_by[name]))
        print('WORKBOOK', [{k:r.get(k) for k in ['Meeting ID','Meeting Name','Day','Start','Attendance','Region','Meeting Page URL']} for r in wb_by[name]])
        print('LIVE', [{k:r.get(k) for k in ['id','name','day','time','attendance_option','region','url','data_source_name','source_slug']} for r in live_by[name]])
