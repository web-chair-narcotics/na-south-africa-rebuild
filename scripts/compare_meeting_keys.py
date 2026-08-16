from collections import Counter
from pathlib import Path
import json
import openpyxl

wb = openpyxl.load_workbook('/home/ubuntu/upload/pasted_file_3nDiNL_NA-South-Africa-All-Meetings.xlsx', read_only=True, data_only=True)
ws = wb['All Meetings']
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
rows = [dict(zip(headers, vals)) for vals in ws.iter_rows(min_row=2, values_only=True)]
report = json.loads(Path('/home/ubuntu/na-south-africa-rebuild/LIVE_PARITY_AUDIT_REPORT.json').read_text())
live = report['liveUniqueRows']

def n(v):
    return '' if v is None else ' '.join(str(v).strip().lower().split())

def key_row(row):
    return (n(row.get('Meeting Name') or row.get('name')), n(row.get('Day') or row.get('day')), n(row.get('Start') or row.get('time')), n(row.get('Region') or row.get('region')))

def url_slug(value):
    if not value: return ''
    return str(value).rstrip('/').split('/')[-1]

keys = {
  'meeting_url': Counter(n(r.get('Meeting Page URL')) for r in rows if r.get('Meeting Page URL')),
  'source_slug': Counter(n(r.get('Source Slug')) for r in rows if r.get('Source Slug')),
  'page_slug': Counter(url_slug(r.get('Meeting Page URL')) for r in rows if r.get('Meeting Page URL')),
  'id': Counter(str(r.get('Meeting ID')) for r in rows if r.get('Meeting ID')),
  'name_day_time_region': Counter(key_row(r) for r in rows),
}
live_keys = {
  'meeting_url': Counter(n(r.get('url')) for r in live if r.get('url')),
  'source_slug': Counter(n(r.get('source_slug')) for r in live if r.get('source_slug')),
  'page_slug': Counter(url_slug(r.get('url')) for r in live if r.get('url')),
  'id': Counter(str(r.get('id')) for r in live if r.get('id')),
  'name_day_time_region': Counter((n(r.get('name')), n(r.get('day')), n(r.get('time')), n(r.get('region'))) for r in live),
}
for name in keys:
    a, b = set(keys[name]), set(live_keys[name])
    print(name, 'workbook_rows', sum(keys[name].values()), 'workbook_unique', len(a), 'live_rows', sum(live_keys[name].values()), 'live_unique', len(b), 'intersection', len(a & b), 'wb_only', len(a-b), 'live_only', len(b-a), 'wb_dupes', sum(c-1 for c in keys[name].values() if c>1), 'live_dupes', sum(c-1 for c in live_keys[name].values() if c>1))
print('Workbook sample only URL slugs:', sorted(set(keys['page_slug']) - set(live_keys['page_slug']))[:20])
print('Live sample only URL slugs:', sorted(set(live_keys['page_slug']) - set(keys['page_slug']))[:20])
